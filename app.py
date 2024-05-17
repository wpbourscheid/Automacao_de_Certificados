import openpyxl
from PIL import Image, ImageDraw, ImageFont
from pathlib import os
import shutil
import PySimpleGUI as sg

# FUNÇÃO PARA VERIFICAR SE A PASTA EXISTE, SE NÃO, CRIA
def Verifica_Dir(newpath):
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        print('Pasta criada')
    else:
        print('Pasta já existe')

# TEMA DAS JANELAS
sg.theme('Green')


janela_principal = [
    [sg.FileBrowse('Escolher Planilha', target='caminho_planilha'), sg.Input(key='caminho_planilha')],
    [sg.FolderBrowse('Salvar em: ', target='caminho_salvamento'), sg.Input(key='caminho_salvamento')],
    [sg.Checkbox('Compactar o arquivo final?', default=True, key='zipar')],
    [sg.CloseButton('Confirmar', key='comecar')]
]

janela = sg.Window('Principal', layout=janela_principal)

# COLETA DE CAMINHOS PELA JANELA
while True:
    event, values = janela.read()
    if event == sg.WIN_CLOSED:
        print('Janela fechada')
        break
    elif event == 'comecar':
        caminho_planilha = values['caminho_planilha']
        caminho_salvamento = values['caminho_salvamento']
        zip_destino = caminho_salvamento
        caminho_salvamento += '/Certificados Salvos'
        compactar = values ['zipar']
        
Verifica_Dir(caminho_salvamento)

#Carregar planilha
workbook_alunos = openpyxl.load_workbook(caminho_planilha)
sheet_alunos = workbook_alunos['Sheet1'] #ESCOLHE A PÁGINA DA PLANILHA 
row_count = sheet_alunos.max_row

#### EDITAR SE PREFERIR OUTRA FONTE #####
fonte_name = ImageFont.truetype('ARIALBD.TTF',90) #FONTE DO NOME
fonte_geral = ImageFont.truetype('ARIAL.TTF',80) #FONTE DAS INFORMAÇÕES
fonte_data = ImageFont.truetype('ARIAL.TTF', 55) #FONTE DAS DATAS

"""TELA DE CARREGAMENTO"""
layout = [[sg.ProgressBar(row_count, key='-PROGRESS_BAR-'), sg.Button('Começar!', key='-UPDATE-')]]
count = 0

window = sg.Window(title='Barra de Progresso', layout=layout, resizable=True)

"""CARREGA AS INFORMAÇÕES DA PLANILHA E INSERE NA IMAGEM ENQUANTO INCREMENTA A BARRA DE LOAD"""
while True:
    event, values = window.read()
    for indice, linha in enumerate(sheet_alunos.iter_rows(min_row=2)):
        nome_curso = linha[0].value #Nome do curso
        nome_participante = linha[1].value # Nome Participante
        tipo_participacao = linha[2].value # Tipo de participação
        data_inicio = linha[3].value # Data Inicio
        data_final = linha[4].value # Data Final
        carga_horaria = linha[5].value # Carga Horária
        data_emissao = linha[6].value # Data emissão certificado

        #Transferir dados da palinha para certificados e salvar
        image = Image.open('certificado_padrao.jpg')
        desenhar = ImageDraw.Draw(image)
        desenhar.text((1020,827), nome_participante, fill='black', font=fonte_name)
        desenhar.text((1060,953), nome_curso, fill= 'black',font= fonte_geral)
        desenhar.text((1430,1068), tipo_participacao, fill= 'black',font= fonte_geral)
        desenhar.text((1480,1188), str(carga_horaria), fill= 'black', font=fonte_geral)
        desenhar.text((750,1770), data_inicio, fill= 'black', font=fonte_data)
        desenhar.text((750,1930), data_final, fill= 'black', font=fonte_data)
        desenhar.text((2220,1930), data_emissao, fill= 'black', font=fonte_data)
        image.save(f'{caminho_salvamento}/{indice} {nome_participante} Certificado.png')

        print(f'Certificado: {indice} salvo')

        #event, values = window.read()

        if event == sg.WIN_CLOSED:
            break
        elif count != row_count-1:
            count += 1
            window['-PROGRESS_BAR-'].update(current_count=count)
        else:
            break
    break

# SE A CHECKBOX FOI MARCADA, COMPACTA O ARQUIVO
if compactar == True:
    shutil.make_archive('Certificados Compactados', 'zip', caminho_salvamento)
    shutil.move("./Certificados Compactados.zip", zip_destino + "\Certificados Compactados.zip")
    print(f'Pasta compactada salva em: {zip_destino}')
else:
    print('Compactação não escolhida')